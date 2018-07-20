# -*- coding: utf-8 -*-
# @Author   : FELIX
# @Date     : 2018/6/30 16:52

import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl


class OperationExcel(object):
    def __init__(self, filename):
        try:
            self.__workbook = xlrd.open_workbook(filename)
        except Exception as e:
            self.creatwb(filename)
            self.__workbook = xlrd.open_workbook(filename)
        self.__wb = load_workbook(filename)

    def get_sheet_names(self):

        return self.__wb.get_sheet_names()

    def get_data(self, sheet_name, row, col, flag='0'):
        """
        :param sheet_name: 表名
        :param row: 行位置
        :param col: 列位置
        :param flag:标志，0：表示获取某行某列，1：获取某行，2：获取某列
        :return:返回
        """
        sheet = self.__workbook.sheet_by_name(sheet_name)
        c = sheet.ncols  # 最大列数
        w = sheet.nrows  # 最大行数
        print(c, w)
        try:
            if str(flag) == '0':
                return sheet.row_values(row - 1)[col - 1]
            if str(flag) == '1':
                return sheet.row_values(row - 1)
            if str(flag) == '2':
                return sheet.col_values(col - 1)
        except Exception as e:
            print('该位置没有数据', e)

    def get_all_data(self, sheet_name):
        '''# 如果下面那种不太好使用注释的这些
        # 获取特定的worksheet
        ws = self.__wb.get_sheet_by_name(sheet_name)
        # 获取表格所有行和列，两者都是可迭代的
        rows = ws.rows
        columns = ws.columns

        # 迭代所有的行
        for row in rows:
            line = [col.value for col in row]
            print(line)
        '''

        sheet = self.__workbook.sheet_by_name(sheet_name)
        datas = []
        data = []
        for i in range(sheet.nrows):
            data.append(sheet.row_values(i))
            # datas.append(data)
        return data

    def get_max_row_and_col(self):
        ws = self.__wb.active
        # sheet = self.__workbook.sheet_by_name(sheet_name)
        # return {'max_rows': sheet.nrows, 'max_cols': sheet.ncols}
        return {'max_rows': ws.max_row, 'max_cols': ws.max_column}

    # 新建excel
    def creatwb(self, filename):
        wb = openpyxl.Workbook()
        wb.save(filename=filename)
        print("新建Excel：" + filename + "成功")

    def write_data(self, filename, col, row, data):
        self.__wb = load_workbook(filename)
        ws = self.__wb.active
        # ws.cell(row=row, column=col).value = data
        ws.cell(row, col, data)
        self.__wb.save(filename)
        self.__wb.close()


if __name__ == '__main__':
    op = OperationExcel('a.xlsx')
    d = op.get_sheet_names()
    print(d)
    print(op.get_data('Sheet2', 1, 2, 0))
    print(op.get_all_data('Sheet1'))
    op.write_data('wass.xlsx', 1, 3, 'test')
    op.write_data('wass.xlsx', 2, 3, 't2est')
